#!/usr/bin/env python3
"""Fail-closed acceptance harness for the TEMP-ONLY Wowcar revenue relabel.

The live repository is the pinned baseline. Two independent temporary copies are
created; the eight approved edits are applied only to the candidate copy.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import importlib.metadata
import json
import math
import os
import re
import shutil
import stat
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

EXPECTED_PYTHON = (3, 13, 5)
EXPECTED_LIBREOFFICE = "25.2.3.2"
EXPECTED_DISTRIBUTIONS = {
    "numpy": "2.2.4",
    "numpy-financial": "1.0.0",
    "openpyxl": "3.1.5",
    "et-xmlfile": "2.0.0",
}
BASELINE_FILES = [
    {
        "mode": "0700",
        "path": "projects/wowcar/generator-source/kod/._belge_uretim",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._build_xlsx.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._capraz.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._kohort.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._model.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0700",
        "path": "projects/wowcar/generator-source/kod/._office",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._opex.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._opex_dokuman.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._recalc.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._sigorta.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._tahsilat.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._tanila.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/._test.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0644",
        "path": "projects/wowcar/generator-source/kod/Wowcar_Finansal_Model_v4.xlsx",
        "sha256": "c50ff41074c7ac02bb3919af72a4c061086bb64e724a0be9052abd78cc8e07a9",
        "size": 143224,
        "type": "file"
    },
    {
        "mode": "0755",
        "path": "projects/wowcar/generator-source/kod/__pycache__",
        "size": 0,
        "type": "dir"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/__pycache__/model.cpython-313.pyc",
        "sha256": "e1ac4074631e91fa49c96aacb29ec7fa7a0de4d6ddeddb757870ae659999417f",
        "size": 24029,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/__pycache__/opex.cpython-313.pyc",
        "sha256": "67b0c56e459e9ed78229b510babd02934f8169b72479b4fe716ca78119982e71",
        "size": 4713,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/__pycache__/sigorta.cpython-313.pyc",
        "sha256": "a9fe0ce2f61d66bfde125859ae7a935f5bb9dcf9c7ab454f1e215d17f5dfa56d",
        "size": 3822,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/__pycache__/tahsilat.cpython-313.pyc",
        "sha256": "347534287f87383b91281b4243fcae3f4b0bf8689f3810a0b510abb3d46b05fa",
        "size": 5029,
        "type": "file"
    },
    {
        "mode": "0700",
        "path": "projects/wowcar/generator-source/kod/belge_uretim",
        "size": 0,
        "type": "dir"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/belge_uretim/._md2pdf.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/belge_uretim/._md2word.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/belge_uretim/md2pdf.py",
        "sha256": "1ff96302f925901da4f48fa4585f5a0727849b2b27a1b1f016e4231bb10a0373",
        "size": 3705,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/belge_uretim/md2word.py",
        "sha256": "cd284a412f39b208de1074a5d5e7da1c0ea65598971523d7f62b0d1f4e9e3214",
        "size": 2236,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/build_xlsx.py",
        "sha256": "b7ae359b1b8f0400e020ed9d59ba718283d4aebef6de39803093d2665f5b8937",
        "size": 37487,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/capraz.py",
        "sha256": "b16a837631d2b6d0d4874a7adb8c9e433002803bbd7e80110a286c12aeded40e",
        "size": 2255,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/kohort.py",
        "sha256": "d4be51cf5f085bd337fbd98e1687b5fe441de456cbe7c9c5e775af658372b455",
        "size": 6359,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/model.py",
        "sha256": "82d96ab6f30053a0a055754ef24fbe18d32a40919d1857ef10a1311d903aa480",
        "size": 20065,
        "type": "file"
    },
    {
        "mode": "0700",
        "path": "projects/wowcar/generator-source/kod/office",
        "size": 0,
        "type": "dir"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/office/._soffice.py",
        "sha256": "ee0487d0ca7428764db1e209fd8f7065c536c90c72c5bdcd17e376c6bd42aaef",
        "size": 163,
        "type": "file"
    },
    {
        "mode": "0755",
        "path": "projects/wowcar/generator-source/kod/office/__pycache__",
        "size": 0,
        "type": "dir"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/office/__pycache__/soffice.cpython-313.pyc",
        "sha256": "8276bcbde94bada29b1f8fb7e3477eab5c9361b38e43eb45839d134e2c2faf90",
        "size": 7799,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/office/soffice.py",
        "sha256": "df2c8d7249c132a4512d744482e4b5ed0ecbe885f0bcfb1295e7e6a7ced0432c",
        "size": 5927,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/opex.py",
        "sha256": "722d86629a12e1a6fd93cb527a988c9585ff10eec9c2c017b3c86440d723b1c6",
        "size": 5836,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/opex_dokuman.py",
        "sha256": "306535987dd983fc6029ca8d7406144dc6c637f5236835bd9684a845e08a34a8",
        "size": 4244,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/recalc.py",
        "sha256": "288cc0b86b15f012f4928ab56e532f8cb11a1cc1f9e0ddb686295b3e36f50d03",
        "size": 11629,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/sigorta.py",
        "sha256": "2d6ea4e3a871b6d3961810fde3d66fb9586ce5536641d4d641d214f0be16e87f",
        "size": 3482,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/tahsilat.py",
        "sha256": "086c91bbce6808b14df7a3e4f4617558dbe925caaae1fef27ee84e70f8a050f8",
        "size": 3547,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/tanila.py",
        "sha256": "50a0a887066c2a56b0b6229d3d9f9bb8304f61cb29c5db0437337e006398dfb6",
        "size": 3848,
        "type": "file"
    },
    {
        "mode": "0600",
        "path": "projects/wowcar/generator-source/kod/test.py",
        "sha256": "3081230503ffe450ba00ed1373d8b441813081acdae63a50ee3ca76ce49a196f",
        "size": 14284,
        "type": "file"
    }
]
BASELINE_ROOT = "8f5cd490bbe6bf4f822f44458a08975864609cb313b05bc97d6a8a394971da6e"

# Ordered longest-form-first within each file. Counts are authoritative.
VOCABULARY_EDITS = {
    "build_xlsx.py": [("Faiz geliri", "Vade farkı geliri", 1)],
    "tanila.py": [("Faiz geliri", "Vade farkı geliri", 2)],
    "kohort.py": [("Faiz geliri", "Vade farkı geliri", 1)],
    "model.py": [("faiz geliri", "vade farkı geliri", 1)],
    "test.py": [
        ("faiz gelirini", "vade farkı gelirini", 1),
        ("faiz geliri", "vade farkı geliri", 1),
        ("Faiz geliri", "Vade farkı geliri", 1),
    ],
}
WIDTH_EDITS = {
    "tanila.py": [
        ("{'Tahsilat':>12}{'Faiz geliri':>12}{'Yatırımcı':>12}", "{'Tahsilat':>12}{'Faiz geliri':>18}{'Yatırımcı':>12}", 1),
        ("{M(v['tahsilat']):>12}{M(v['faiz']):>12}{M(v['yat']):>12}", "{M(v['tahsilat']):>12}{M(v['faiz']):>18}{M(v['yat']):>12}", 1),
        ("{'Yatırımcı oranı':>17}{'Faiz geliri':>14}{'Yatırımcı':>13}", "{'Yatırımcı oranı':>17}{'Faiz geliri':>18}{'Yatırımcı':>13}", 1),
        ("{'%'+n(r['yat_oran']*100,0):>17}{M(v['faiz']):>14}{M(v['yat']):>13}", "{'%'+n(r['yat_oran']*100,0):>17}{M(v['faiz']):>18}{M(v['yat']):>13}", 1),
    ],
    "kohort.py": [
        ("{'Faiz geliri':>13}", "{'Faiz geliri':>18}", 1),
        ("{M(sum(r['faiz'] for r in rr)):>13}", "{M(sum(r['faiz'] for r in rr)):>18}", 1),
    ],
}
# Independent stdout schema. This must not be derived from WIDTH_EDITS.
STDOUT_TABLES = {
    "tanila.py": [
        {"marker": "A. YILLIK ÖZET", "start": 28, "old_width": 12, "target_width": 18, "rows": 5},
        {"marker": "D. RESET", "start": 36, "old_width": 14, "target_width": 18, "rows": 5},
    ],
    "kohort.py": [
        {"marker": "Mod ", "start": 17, "old_width": 13, "target_width": 18, "rows": 5},
    ],
}
SCENARIOS = [
    ("default", {}),
    ("reset", {"reset": True}),
    ("gecikme_1", {"gecikme": 1}),
    ("gecikme_2", {"gecikme": 2}),
    ("kapasite_400", {"kapasite": 400}),
    ("kapasite_800", {"kapasite": 800}),
    ("musteri_090", {"musteri": 0.90}),
    ("musteri_085", {"musteri": 0.85}),
    ("mod_B", {"mod": "B"}),
]
PROBE = r'''
import json, math, model
def c(v):
    if v is None: return ["none"]
    if type(v) is bool: return ["bool", v]
    if type(v) is int: return ["int", str(v)]
    if type(v) is float:
        if not math.isfinite(v): raise ValueError("non-finite numeric leaf")
        return ["float", v.hex()]
    if type(v) is str: return ["str", v]
    if type(v) is list: return ["list", [c(x) for x in v]]
    if type(v) is tuple: return ["tuple", [c(x) for x in v]]
    if type(v) is dict:
        return ["dict", [[c(k), c(v[k])] for k in sorted(v, key=lambda x:(type(x).__name__, repr(x)))]]
    raise TypeError("unsupported leaf: " + type(v).__name__)
scenarios = json.loads(%r)
out = []
for name, kw in scenarios:
    log, params = model.calis(**kw)
    out.append([name, c(log), c(params), c(model.yillik(log))])
print(json.dumps(out, ensure_ascii=False, separators=(",", ":"), sort_keys=True))
'''


class AcceptanceFailure(RuntimeError):
    pass


class ConfigurationFailure(RuntimeError):
    pass


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def canonical_digest(value: Any) -> str:
    data = json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True).encode()
    return sha256_bytes(data)


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AcceptanceFailure(message)


def config_require(condition: bool, message: str) -> None:
    if not condition:
        raise ConfigurationFailure(message)


def find_repo() -> Path:
    repo = Path(__file__).resolve().parents[2]
    require((repo / "projects/wowcar/generator-source/kod").is_dir(), "repository layout mismatch")
    return repo


def verify_runtime() -> dict[str, str]:
    config_require(sys.implementation.name == "cpython", "CPython is required")
    config_require(sys.version_info[:3] == EXPECTED_PYTHON, f"CPython {EXPECTED_PYTHON!r} is required")
    versions = {name: importlib.metadata.version(name) for name in EXPECTED_DISTRIBUTIONS}
    config_require(versions == EXPECTED_DISTRIBUTIONS, f"dependency mismatch: {versions!r}")
    proc = subprocess.run(["libreoffice", "--version"], capture_output=True, text=True, timeout=10)
    config_require(proc.returncode == 0, "LibreOffice version probe failed")
    match = re.search(r"LibreOffice\s+([^\s]+)", proc.stdout)
    config_require(bool(match) and match.group(1) == EXPECTED_LIBREOFFICE, "LibreOffice version mismatch")
    return {"python": ".".join(map(str, sys.version_info[:3])), "libreoffice": match.group(1), **versions}


def baseline_records(kod: Path, repo: Path) -> list[dict[str, Any]]:
    records = []
    for path in sorted(kod.rglob("*")):
        st = path.lstat()
        record: dict[str, Any] = {
            "path": path.relative_to(repo).as_posix(),
            "mode": format(stat.S_IMODE(st.st_mode), "04o"),
        }
        if path.is_symlink() or not (path.is_file() or path.is_dir()):
            raise AcceptanceFailure(f"baseline contains forbidden member type: {record['path']}")
        if path.is_file():
            data = path.read_bytes()
            record.update(type="file", size=len(data), sha256=sha256_bytes(data))
        else:
            record.update(type="dir", size=0)
        records.append(record)
    return records


def verify_baseline(kod: Path, repo: Path) -> None:
    actual = baseline_records(kod, repo)
    config_require(actual == BASELINE_FILES, "pinned complete-tree baseline manifest mismatch")
    config_require(canonical_digest(actual) == BASELINE_ROOT, "embedded baseline root digest mismatch")


def tree_manifest(paths: list[tuple[str, Path]]) -> tuple[str, list[dict[str, Any]]]:
    records: list[dict[str, Any]] = []
    for label, root in paths:
        require(root.exists(), f"protected path missing: {root}")
        members = [root] if root.is_file() else sorted(root.rglob("*"))
        for member in members:
            rel = "." if member == root else member.relative_to(root).as_posix()
            st = member.lstat()
            record: dict[str, Any] = {"root": label, "path": rel, "mode": stat.S_IMODE(st.st_mode)}
            if member.is_symlink():
                record.update(type="symlink", target=os.readlink(member))
            elif member.is_file():
                data = member.read_bytes()
                record.update(type="file", size=len(data), sha256=sha256_bytes(data))
            elif member.is_dir():
                record.update(type="dir")
            else:
                record.update(type="other")
            records.append(record)
    return canonical_digest(records), records


def inventory(root: Path) -> dict[str, tuple[str, int, str]]:
    out: dict[str, tuple[str, int, str]] = {}
    for path in sorted(root.rglob("*")):
        rel = path.relative_to(root).as_posix()
        st = path.lstat()
        if path.is_symlink():
            out[rel] = ("symlink", stat.S_IMODE(st.st_mode), os.readlink(path))
        elif path.is_file():
            out[rel] = ("file", stat.S_IMODE(st.st_mode), sha256_bytes(path.read_bytes()))
        elif path.is_dir():
            out[rel] = ("dir", stat.S_IMODE(st.st_mode), "")
        else:
            out[rel] = ("other", stat.S_IMODE(st.st_mode), "")
    return out


def apply_candidate_edits(root: Path) -> None:
    for edits in (WIDTH_EDITS, VOCABULARY_EDITS):
      for rel, replacements in edits.items():
        path = root / rel
        text = path.read_text(encoding="utf-8")
        for old, new, count in replacements:
            require(text.count(old) == count, f"preimage count mismatch: {rel}: {old!r}")
            text = text.replace(old, new)
        path.write_text(text, encoding="utf-8", newline="")


def check_source_boundary(baseline: Path, candidate: Path, mode: str) -> dict[str, Any]:
    before, after = inventory(baseline), inventory(candidate)
    require(before.keys() == after.keys(), "added, removed, or renamed candidate member")
    changed = sorted(path for path in before if before[path] != after[path])
    expected_changed = [] if mode == "unchanged" else sorted(VOCABULARY_EDITS)
    require(changed == expected_changed, f"changed-file boundary mismatch: {changed!r}")
    if mode == "unchanged":
        return {"changed_files": [], "vocabulary_anchors": 0, "width_anchors": 0, "controlled_edits": 0}
    for rel in changed:
        require(before[rel][0] == after[rel][0] == "file", f"non-regular approved member: {rel}")
        normalized = (candidate / rel).read_text(encoding="utf-8")
        for old, new, count in VOCABULARY_EDITS[rel]:
            require(normalized.count(new) == count, f"postimage count mismatch: {rel}: {new!r}")
            normalized = normalized.replace(new, old)
        for old, new, count in WIDTH_EDITS.get(rel, []):
            # Annual/reset header edits intentionally converge on the same width-18
            # postimage. Reverse them deterministically in source order; the final
            # byte-equality check proves the mapping and ordering are exact.
            require(normalized.count(new) >= count, f"width postimage count mismatch: {rel}: {new!r}")
            normalized = normalized.replace(new, old, count)
        require(normalized.encode() == (baseline / rel).read_bytes(), f"normalized byte mismatch: {rel}")
    forms = {"title": 0, "lower": 0, "inflected": 0}
    per_file = {}
    for rel, replacements in VOCABULARY_EDITS.items():
        per_file[rel] = sum(count for _, _, count in replacements)
        for _, new, count in replacements:
            forms["inflected" if new == "vade farkı gelirini" else "title" if new[0].isupper() else "lower"] += count
    require(forms == {"title": 5, "lower": 2, "inflected": 1}, f"form distribution mismatch: {forms}")
    require(per_file == {"build_xlsx.py": 1, "tanila.py": 2, "kohort.py": 1, "model.py": 1, "test.py": 3}, "file distribution mismatch")
    return {"changed_files": changed, "forms": forms, "vocabulary_anchors": 8, "width_anchors": 6, "controlled_edits": 14}


def run(root: Path, argv: list[str], timeout: int = 180) -> dict[str, Any]:
    tmp = root / ".acceptance-tmp"
    tmp.mkdir(exist_ok=True)
    env = os.environ.copy()
    env.update(PYTHONDONTWRITEBYTECODE="1", PYTHONHASHSEED="0", TMPDIR=str(tmp))
    proc = subprocess.run(argv, cwd=root, env=env, capture_output=True, timeout=timeout)
    return {"argv": argv, "returncode": proc.returncode, "stdout": proc.stdout, "stderr": proc.stderr}


def require_success(result: dict[str, Any], label: str) -> None:
    require(result["returncode"] == 0, f"{label} exited {result['returncode']}: {result['stderr'][-400:]!r}")


def numeric_probe(root: Path) -> tuple[Any, dict[str, Any]]:
    code = PROBE % json.dumps(SCENARIOS, ensure_ascii=False)
    result = run(root, [sys.executable, "-c", code])
    require_success(result, "numeric probe")
    return json.loads(result["stdout"]), result


def numeric_leaf_count(value: Any) -> int:
    if isinstance(value, list) and len(value) == 2 and value[0] in ("int", "float"):
        return 1
    if isinstance(value, (list, dict)):
        children = value if isinstance(value, list) else value.values()
        return sum(numeric_leaf_count(item) for item in children)
    return 0


def scrub_command(result: dict[str, Any]) -> dict[str, Any]:
    return {"argv": result["argv"], "exit_status": result["returncode"]}


def build_chain(root: Path) -> tuple[Path, Path, dict[str, dict[str, Any]]]:
    workbook = root / "Wowcar_Finansal_Model_v4.xlsx"
    commands = {}
    commands["build"] = run(root, [sys.executable, "build_xlsx.py"])
    require_success(commands["build"], "build_xlsx.py")
    require(workbook.is_file(), "build did not produce workbook")
    pre = root / "Wowcar_Finansal_Model_v4.pre-recalc.xlsx"
    shutil.copy2(workbook, pre)
    commands["recalc"] = run(root, [sys.executable, "recalc.py", workbook.name])
    require_success(commands["recalc"], "recalc.py")
    commands["test"] = run(root, [sys.executable, "test.py"])
    require_success(commands["test"], "test.py")
    test_summary = re.search(rb"DO\xc4\x9eRULAMA:\s+(\d+)/(\d+)\s+test ge\xc3\xa7ti", commands["test"]["stdout"])
    require(bool(test_summary), "test.py summary is missing")
    require(test_summary.group(1) == test_summary.group(2) and b"\xe2\x9c\x97" not in commands["test"]["stdout"], "test.py reported a failing predicate")
    commands["capraz"] = run(root, [sys.executable, "capraz.py"])
    require_success(commands["capraz"], "capraz.py")
    require("TÜM SÜTUNLAR UYUMLU".encode() in commands["capraz"]["stdout"], "capraz.py reported a reconciliation failure")
    return pre, workbook, commands


def scalar(value: Any) -> Any:
    import datetime
    from openpyxl.worksheet.formula import ArrayFormula

    if isinstance(value, ArrayFormula):
        return ["array_formula", value.text, value.ref]
    if value is None: return ["blank"]
    if type(value) is bool: return ["bool", value]
    if type(value) is int: return ["int", str(value)]
    if type(value) is float:
        require(math.isfinite(value), "non-finite workbook numeric leaf")
        return ["float", value.hex()]
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time, datetime.timedelta)):
        return [type(value).__name__, value.isoformat() if hasattr(value, "isoformat") else str(value)]
    if isinstance(value, str): return ["string", value]
    return [type(value).__name__, repr(value)]


def workbook_structure(path: Path) -> dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        sheets = []
        for ws in wb.worksheets:
            cells = []
            for (row, col), cell in sorted(ws._cells.items()):
                hyperlink = None if cell.hyperlink is None else [cell.hyperlink.target, cell.hyperlink.location, cell.hyperlink.display, cell.hyperlink.tooltip]
                comment = None if cell.comment is None else [cell.comment.text, cell.comment.author, cell.comment.width, cell.comment.height]
                cells.append([cell.coordinate, cell.data_type, scalar(cell.value), cell.number_format, comment, hyperlink, list(cell._style) if cell.has_style else None])
            row_dims = [[str(k), str(v)] for k, v in sorted(ws.row_dimensions.items())]
            col_dims = [[str(k), str(v)] for k, v in sorted(ws.column_dimensions.items())]
            validations = [str(v) for v in ws.data_validations.dataValidation]
            conditional = [[str(k), [str(rule) for rule in rules]] for k, rules in ws.conditional_formatting._cf_rules.items()]
            tables = [[name, str(table)] for name, table in sorted(ws.tables.items())]
            sheets.append({
                "title": ws.title, "cells": cells, "merged": sorted(map(str, ws.merged_cells.ranges)),
                "freeze": str(ws.freeze_panes or ""), "row_dimensions": row_dims, "column_dimensions": col_dims,
                "validations": validations, "conditional": conditional, "tables": tables,
                "auto_filter": (ws.auto_filter.ref, sorted(map(str, ws.auto_filter.filterColumn)), str(ws.auto_filter.sortState) if ws.auto_filter.sortState else None), "print_area": str(ws.print_area), "print_title_rows": str(ws.print_title_rows),
                "print_title_cols": str(ws.print_title_cols), "sheet_properties": str(ws.sheet_properties),
                "page_margins": str(ws.page_margins), "page_setup": str(ws.page_setup), "print_options": str(ws.print_options),
            })
        names = [[name, str(value)] for name, value in wb.defined_names.items()]
        return {"sheetnames": wb.sheetnames, "defined_names": names, "sheets": sheets}
    finally:
        wb.close()


def cached_cells(path: Path) -> tuple[dict[str, Any], int]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    count = 0
    out = {"sheetnames": wb.sheetnames, "sheets": []}
    try:
        for ws in wb.worksheets:
            records = []
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row, col)
                    encoded = scalar(cell.value)
                    if encoded[0] in ("int", "float"): count += 1
                    records.append([cell.coordinate, encoded])
            out["sheets"].append([ws.title, records])
        return out, count
    finally:
        wb.close()


def normalize_f4_pair(baseline: dict[str, Any], candidate: dict[str, Any], cached: bool = False) -> None:
    def locate(data: dict[str, Any]) -> list[Any]:
        if cached:
            sheet = next(x for x in data["sheets"] if x[0] == "Aylık Motor")
            return next(x for x in sheet[1] if x[0] == "F4")
        sheet = next(x for x in data["sheets"] if x["title"] == "Aylık Motor")
        return next(x for x in sheet["cells"] if x[0] == "F4")
    old, new = locate(baseline), locate(candidate)
    value_index = 1 if cached else 2
    require(old[value_index] == ["string", "Faiz geliri"], "baseline Aylık Motor!F4 mismatch")
    require(new[value_index] == ["string", "Vade farkı geliri"], "candidate Aylık Motor!F4 mismatch")
    new[value_index] = old[value_index]


def compare_workbooks(base_path: Path, candidate_path: Path, cached: bool, mode: str) -> tuple[int, int]:
    if cached:
        left, left_count = cached_cells(base_path)
        right, right_count = cached_cells(candidate_path)
    else:
        left, right = workbook_structure(base_path), workbook_structure(candidate_path)
        left_count = right_count = sum(len(s["cells"]) for s in left["sheets"])
    require(left_count == right_count, "workbook record-count mismatch")
    if mode == "14-anchor":
        normalize_f4_pair(left, right, cached=cached)
    require(left == right, "workbook structure/value mismatch outside the mode-owned boundary")
    return left_count, right_count


def normalize_stdout(data: bytes, rel: str, mode: str) -> bytes:
    if mode == "unchanged":
        return data
    text = data.decode("utf-8")
    if rel == "test.py":
        for old, new, _ in VOCABULARY_EDITS[rel]:
            text = text.replace(old, new)
        return text.encode("utf-8")
    lines = text.splitlines(keepends=True)
    for schema in STDOUT_TABLES[rel]:
        matches = [i for i, line in enumerate(lines) if schema["marker"] in line]
        require(len(matches) == 1, f"C6 table marker mismatch: {rel}: {schema['marker']!r}")
        header_index = matches[0] + 2
        require(header_index + schema["rows"] < len(lines), f"C6 table truncated: {rel}")
        start, old_width, target_width = schema["start"], schema["old_width"], schema["target_width"]
        old_header = "Faiz geliri".rjust(old_width)
        require(lines[header_index][start:start + old_width] == old_header, f"C6 old header field mismatch: {rel}")
        lines[header_index] = lines[header_index][:start] + " Vade farkı geliri" + lines[header_index][start + old_width:]
        for row_index in range(header_index + 1, header_index + 1 + schema["rows"]):
            field = lines[row_index][start:start + old_width]
            require(len(field) == old_width and field.strip(), f"C6 numeric field mismatch: {rel}")
            lines[row_index] = lines[row_index][:start] + (" " * (target_width - old_width)) + field + lines[row_index][start + old_width:]
    return "".join(lines).encode("utf-8")


def compare_stdout(base: Path, candidate: Path, base_chain: dict[str, Any], candidate_chain: dict[str, Any], mode: str) -> dict[str, Any]:
    compared = {}
    for rel in ("tanila.py", "kohort.py"):
        left, right = run(base, [sys.executable, rel]), run(candidate, [sys.executable, rel])
        errors = []
        if left["returncode"] != right["returncode"] or left["returncode"] != 0: errors.append("exit status")
        if left["stderr"] != right["stderr"]: errors.append("stderr")
        try:
            if normalize_stdout(left["stdout"], rel, mode) != right["stdout"]: errors.append("stdout")
        except AcceptanceFailure as exc:
            errors.append(str(exc))
        compared[rel] = {"result": "FAIL" if errors else "PASS", "errors": errors}
    left, right = base_chain["test"], candidate_chain["test"]
    errors = []
    if left["returncode"] != right["returncode"] or left["returncode"] != 0: errors.append("exit status")
    if left["stderr"] != right["stderr"]: errors.append("stderr")
    if normalize_stdout(left["stdout"], "test.py", mode) != right["stdout"]: errors.append("stdout")
    compared["test.py"] = {"result": "FAIL" if errors else "PASS", "errors": errors}
    require(all(item["result"] == "PASS" for item in compared.values()), "C6 output comparison failed")
    return compared


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", required=True, choices=("unchanged", "14-anchor"))
    parser.add_argument("--candidate", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    return parser.parse_args()


def control_identity(repo: Path) -> dict[str, str]:
    paths = {
        "harness_sha256": Path(__file__),
        "wrapper_sha256": repo / "tests/test_wowcar_revenue_vocabulary_acceptance.sh",
        "requirements_sha256": repo / "tests/requirements-wowcar-revenue-vocabulary.txt",
    }
    for path in paths.values():
        config_require(path.is_file() and not path.is_symlink(), f"missing control file: {path}")
    identity = {key: sha256_bytes(path.read_bytes()) for key, path in paths.items()}
    commit = subprocess.run(
        ["git", "-C", str(repo), "rev-parse", "HEAD"],
        capture_output=True,
        text=True,
        timeout=10,
    )
    identity["repository_commit"] = commit.stdout.strip() if commit.returncode == 0 else "UNAVAILABLE-NOT-CANONICALLY-LANDED"
    identity["deployment_identity"] = "current-filesystem-bytes"
    return identity


def main() -> int:
    args = parse_args()
    started = dt.datetime.now(dt.timezone.utc)
    report: dict[str, Any] = {
        "schema_version": 2,
        "started_utc": started.isoformat(),
        "mode": args.mode,
        "checks": {},
        "exit_classification": "configuration-error",
    }
    repo = find_repo().resolve()
    live = (repo / "projects/wowcar/generator-source/kod").resolve()
    candidate_arg = args.candidate.resolve()
    report_path = args.report.resolve()
    protected = [("projects/wowcar", repo / "projects/wowcar")]
    protected += [(p, repo / p) for p in ("memories/consensus.md", "memories/operator-requests.md", "memories/human-directive.md")]
    pre_digest = ""
    pre_records: list[dict[str, Any]] = []
    status = 2
    try:
        config_require(args.report.is_absolute(), "report path must be absolute")
        config_require(not report_path.is_relative_to(repo), "report path must be outside repository")
        config_require(candidate_arg == live, "candidate must resolve to the exact approved live kod path")
        report["runtime"] = verify_runtime()
        verify_baseline(live, repo)
        report["baseline_manifest_root"] = BASELINE_ROOT
        report["baseline_member_count"] = len(BASELINE_FILES)
        report["control_identity"] = control_identity(repo)
        live_pre = baseline_records(live, repo)
        report["candidate_pre_root"] = canonical_digest(live_pre)
        pre_digest, pre_records = tree_manifest(protected)
        with tempfile.TemporaryDirectory(prefix="wowcar-revenue-baseline-") as baseline_tmp, tempfile.TemporaryDirectory(prefix="wowcar-revenue-candidate-") as candidate_tmp:
            baseline = Path(baseline_tmp) / "kod"
            candidate = Path(candidate_tmp) / "kod"
            shutil.copytree(live, baseline, symlinks=False)
            shutil.copytree(live, candidate, symlinks=False)
            if args.mode == "14-anchor":
                apply_candidate_edits(candidate)
            report["checks"]["C1"] = {"result": "PASS", **check_source_boundary(baseline, candidate, args.mode)}

            base_numeric, _ = numeric_probe(baseline)
            candidate_numeric, _ = numeric_probe(candidate)
            require(base_numeric == candidate_numeric, "numeric scenario canonicalization mismatch")
            report["checks"]["C2"] = {
                "result": "PASS",
                "scenarios": len(SCENARIOS),
                "numeric_leaves": numeric_leaf_count(base_numeric),
                "max_absolute_delta": 0,
                "max_relative_delta": 0,
            }

            base_pre, base_calc, base_chain = build_chain(baseline)
            candidate_pre, candidate_calc, candidate_chain = build_chain(candidate)
            for key in ("build", "recalc"):
                require(base_chain[key]["stdout"] == candidate_chain[key]["stdout"], f"{key} stdout mismatch")
                require(base_chain[key]["stderr"] == candidate_chain[key]["stderr"], f"{key} stderr mismatch")
            require(base_chain["capraz"]["stdout"] == candidate_chain["capraz"]["stdout"], "capraz.py stdout mismatch")
            require(base_chain["capraz"]["stderr"] == candidate_chain["capraz"]["stderr"], "capraz.py stderr mismatch")
            report["checks"]["C3"] = {
                "result": "PASS",
                "baseline_commands": [scrub_command(v) for v in base_chain.values()],
                "candidate_commands": [scrub_command(v) for v in candidate_chain.values()],
            }

            pre_count, _ = compare_workbooks(base_pre, candidate_pre, cached=False, mode=args.mode)
            calc_count, _ = compare_workbooks(base_calc, candidate_calc, cached=False, mode=args.mode)
            report["checks"]["C4"] = {
                "result": "PASS",
                "pre_recalculation_records": pre_count,
                "recalculated_records": calc_count,
                "allowed_difference": None if args.mode == "unchanged" else "Aylık Motor!F4: Faiz geliri -> Vade farkı geliri",
            }
            cached_count, _ = compare_workbooks(base_calc, candidate_calc, cached=True, mode=args.mode)
            report["checks"]["C5"] = {
                "result": "PASS",
                "cached_records": cached_count,
                "max_absolute_delta": 0,
                "max_relative_delta": 0,
                "allowed_difference": None if args.mode == "unchanged" else "Aylık Motor!F4",
            }
            report["checks"]["C6"] = {
                "result": "PASS",
                "oracle": "direct byte equality" if args.mode == "unchanged" else "independent baseline fixed-column schema",
                "commands": compare_stdout(baseline, candidate, base_chain, candidate_chain, args.mode),
            }
            report["temporary_trees"] = {"baseline": "cleaned-by-context", "candidate": "cleaned-by-context"}

        live_post = baseline_records(live, repo)
        report["candidate_post_root"] = canonical_digest(live_post)
        require(live_pre == live_post, "validated live candidate tree changed during acceptance")
        post_digest, post_records = tree_manifest(protected)
        require(pre_records == post_records and pre_digest == post_digest, "protected surface changed during acceptance")
        report["checks"]["C7"] = {
            "result": "PASS",
            "label": "protected-surface non-mutation",
            "pre_root": pre_digest,
            "post_root": post_digest,
        }
        report["result"] = "PASS"
        report["exit_classification"] = "pass"
        status = 0
    except AcceptanceFailure as exc:
        report["result"] = "FAIL"
        report["failure"] = str(exc)
        report["exit_classification"] = "comparison-failure"
        status = 1
    except (ConfigurationFailure, FileNotFoundError, ImportError, importlib.metadata.PackageNotFoundError, subprocess.TimeoutExpired) as exc:
        report["result"] = "ERROR"
        report["failure"] = f"{type(exc).__name__}: {exc}"
        report["exit_classification"] = "configuration-error"
        status = 2
    except Exception as exc:
        report["result"] = "ERROR"
        report["failure"] = f"unexpected {type(exc).__name__}: {exc}"
        report["exit_classification"] = "unexpected-error"
        status = 2

    if pre_digest and "C7" not in report["checks"]:
        try:
            post_digest, post_records = tree_manifest(protected)
            unchanged = pre_digest == post_digest and pre_records == post_records
            report["checks"]["C7"] = {
                "result": "PASS" if unchanged else "FAIL",
                "label": "protected-surface non-mutation",
                "pre_root": pre_digest,
                "post_root": post_digest,
            }
            if not unchanged:
                report["result"] = "FAIL"
                report["failure"] = "protected surface changed during failed acceptance run"
                report["exit_classification"] = "comparison-failure"
                status = 1
        except Exception as exc:
            report["result"] = "ERROR"
            report["failure"] = f"post-run protected-surface hash failed: {type(exc).__name__}: {exc}"
            report["exit_classification"] = "cleanup-or-verification-error"
            status = 2

    report["finished_utc"] = dt.datetime.now(dt.timezone.utc).isoformat()
    payload = json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    try:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(payload, encoding="utf-8")
    except Exception as exc:
        report["result"] = "ERROR"
        report["failure"] = f"report write failed: {type(exc).__name__}: {exc}"
        report["exit_classification"] = "report-error"
        status = 2
        payload = json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    sys.stdout.write(payload)
    return status


if __name__ == "__main__":
    raise SystemExit(main())
